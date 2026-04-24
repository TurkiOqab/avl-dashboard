from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import BinaryIO, Union

from openpyxl import load_workbook


# Header names as they appear in the source Excel file.
# Update these to match your AVL report (e.g. Arabic labels).
HEADERS = {
    "vehicle_type": "Type",
    "plate_number": "Plate Number",
    "record_date": "Date",
    "visits_count": "Number of Locations Visited",
    "location_indices": "Location Index Number",
}


@dataclass
class ParsedReport:
    report_date: date
    records: list[dict]


def parse_report(source: Union[str, Path, BinaryIO]) -> ParsedReport:
    wb = load_workbook(source, data_only=True, read_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise ValueError("Empty workbook.")

    col_index = _build_column_index(header_row)

    records: list[dict] = []
    for row in rows_iter:
        if _is_empty(row):
            break  # end of data / start of photo section
        records.append({
            "vehicle_type": _str(row[col_index["vehicle_type"]]),
            "plate_number": _str(row[col_index["plate_number"]]),
            "record_date": _as_date(row[col_index["record_date"]]),
            "visits_count": int(row[col_index["visits_count"]] or 0),
            "location_indices": _str(row[col_index["location_indices"]]),
        })

    if not records:
        raise ValueError("No data rows found in the report.")

    report_date = max(r["record_date"] for r in records)
    return ParsedReport(report_date=report_date, records=records)


def _build_column_index(header_row) -> dict:
    normalized = [_normalize(h) for h in header_row]
    idx = {}
    for key, name in HEADERS.items():
        target = _normalize(name)
        if target not in normalized:
            raise ValueError(f"Missing required column: '{name}'")
        idx[key] = normalized.index(target)
    return idx


def _normalize(s) -> str:
    return "" if s is None else str(s).strip().lower()


def _str(v) -> str:
    return "" if v is None else str(v).strip()


def _as_date(v) -> date:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    raise ValueError(f"Cannot interpret {v!r} as a date.")


def _is_empty(row) -> bool:
    return all(c is None or (isinstance(c, str) and c.strip() == "") for c in row)
