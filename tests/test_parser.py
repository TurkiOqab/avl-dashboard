from datetime import date

import pytest
from openpyxl import Workbook, load_workbook

from src.parser import HEADERS, parse_report


def test_parse_report_extracts_all_rows(make_excel):
    rows = [
        {
            "vehicle_type": "Sedan",
            "plate_number": "ABC-1234",
            "record_date": date(2026, 4, 22),
            "visits_count": 3,
            "location_indices": "12, 13, 14",
        },
        {
            "vehicle_type": "Van",
            "plate_number": "XYZ-9999",
            "record_date": date(2026, 4, 22),
            "visits_count": 1,
            "location_indices": "7",
        },
    ]
    path = make_excel(rows)
    parsed = parse_report(path)

    assert parsed.report_date == date(2026, 4, 22)
    assert len(parsed.records) == 2
    assert parsed.records[0]["plate_number"] == "ABC-1234"
    assert parsed.records[0]["visits_count"] == 3
    assert parsed.records[0]["location_indices"] == "12, 13, 14"
    assert parsed.records[0]["record_date"] == date(2026, 4, 22)


def test_missing_header_raises(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["Wrong1", "Wrong2", "Wrong3", "Wrong4", "Wrong5"])
    path = tmp_path / "bad.xlsx"
    wb.save(path)

    with pytest.raises(ValueError, match="Missing required column"):
        parse_report(path)


def test_no_data_rows_raises(make_excel):
    path = make_excel([])
    with pytest.raises(ValueError, match="No data rows"):
        parse_report(path)


def test_empty_row_terminates_parsing(make_excel):
    """An empty row marks end-of-data. Everything below (the photo section)
    must be ignored."""
    rows = [
        {
            "vehicle_type": "Sedan",
            "plate_number": "ABC-1234",
            "record_date": date(2026, 4, 22),
            "visits_count": 2,
            "location_indices": "1, 2",
        }
    ]
    path = make_excel(rows, filename="with-photos.xlsx")

    wb = load_workbook(path)
    ws = wb.active
    ws.append([None, None, None, None, None])
    ws.append(["PHOTO SECTION", None, None, None, None])
    ws.append(["Photo 1", None, None, None, None])
    wb.save(path)

    parsed = parse_report(path)
    assert len(parsed.records) == 1
    assert parsed.records[0]["plate_number"] == "ABC-1234"


def test_parse_report_accepts_iso_string_dates(tmp_path):
    """Some AVL exports write the date column as an ISO string rather than
    an Excel date serial. The parser must accept both."""
    wb = Workbook()
    ws = wb.active
    ws.append([
        HEADERS["vehicle_type"],
        HEADERS["plate_number"],
        HEADERS["record_date"],
        HEADERS["visits_count"],
        HEADERS["location_indices"],
    ])
    ws.append(["Sedan", "ABC-1234", "2026-04-22", 3, "12, 13, 14"])
    path = tmp_path / "string-dates.xlsx"
    wb.save(path)

    parsed = parse_report(path)
    assert parsed.report_date == date(2026, 4, 22)
    assert parsed.records[0]["record_date"] == date(2026, 4, 22)
